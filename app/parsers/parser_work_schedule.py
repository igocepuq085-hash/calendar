from datetime import date
from typing import Any

from openpyxl.utils import get_column_letter

from app.models import ShiftType
from app.parsers.base import ParseError, ParseResult, cell_value_with_merge
from app.services.kip import default_shift_times
from app.services.people import normalize_name, normalize_tab_number


class WorkScheduleParser:
    name = "work_schedule"

    def detect(self, workbook: Any) -> bool:
        for ws in workbook.worksheets:
            title = " ".join(str(cell_value_with_merge(ws, row, 1) or "") for row in range(1, min(ws.max_row, 10) + 1))
            if "Г Р А Ф И К" in title or "работы машинистов" in title:
                return True
        return False

    def parse(self, workbook: Any) -> ParseResult:
        ws = workbook.worksheets[0]
        month, year = 5, 2026
        title = str(cell_value_with_merge(ws, 7, 1) or "")
        if "2026" in title and "Май" in title:
            month, year = 5, 2026

        rows: list[dict[str, Any]] = []
        errors: list[ParseError] = []
        day_columns: list[tuple[int, int]] = []
        for col in range(1, ws.max_column + 1):
            value = ws.cell(10, col).value
            if isinstance(value, int) and 1 <= value <= 31:
                day_columns.append((col, value))

        for row_num in range(11, ws.max_row + 1):
            full_name = ws.cell(row_num, 2).value
            if not full_name:
                continue
            if str(full_name).startswith("Кол-во"):
                break
            tab_number = normalize_tab_number(ws.cell(row_num, 3).value)
            position = ws.cell(row_num, 4).value
            employee_name = normalize_name(full_name)
            for col, day in day_columns:
                raw = ws.cell(row_num, col).value
                if raw in (None, ""):
                    continue
                shift_type = self._shift_type(raw)
                shift_date = date(year, month, day)
                start, end = default_shift_times(shift_date, shift_type)
                rows.append(
                    {
                        "row_number": row_num,
                        "kind": self.name,
                        "full_name": employee_name,
                        "tab_number": tab_number,
                        "position": str(position or ""),
                        "shift_date": shift_date,
                        "shift_type": shift_type,
                        "start_datetime": start,
                        "end_datetime": end,
                        "raw_value": str(raw),
                        "source_cell": f"{get_column_letter(col)}{row_num}",
                    }
                )
        errors.extend(self.validate(rows))
        return ParseResult(parser_name=self.name, rows=rows, errors=errors)

    def validate(self, rows: list[dict[str, Any]]) -> list[ParseError]:
        errors: list[ParseError] = []
        allowed = {item.value for item in ShiftType}
        for row in rows:
            if row["shift_type"].value not in allowed:
                errors.append(ParseError(row["row_number"], "Неизвестный тип смены", row))
            if not row.get("full_name"):
                errors.append(ParseError(row["row_number"], "Не найдено ФИО", row))
        return errors

    def _shift_type(self, raw: Any) -> ShiftType:
        value = str(raw).strip().lower()
        if value in {"в", "выходной"}:
            return ShiftType.off
        if value in {"о", "от", "отпуск"}:
            return ShiftType.vacation
        if value in {"б", "больничный"}:
            return ShiftType.sick
        if value in {"у", "об", "уч"}:
            return ShiftType.training
        try:
            hours = float(value.replace(",", "."))
        except ValueError:
            return ShiftType.unknown
        if hours >= 10:
            return ShiftType.day
        return ShiftType.night
