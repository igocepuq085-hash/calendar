from datetime import datetime

from openpyxl import Workbook

from app.parsers.parser_medical import MedicalParser
from app.parsers.parser_knowledge import KnowledgeParser


def test_screen_parser_reads_all_non_medical_check_types() -> None:
    wb = Workbook()
    ws = wb.active
    ws.cell(3, 6).value = "Фильтр"
    ws.cell(3, 8).value = "Проверка знаний по ОТ"
    ws.cell(3, 10).value = "Проверка знаний по ЭБ"
    ws.cell(3, 12).value = "Проверка знаний Стропальщики"
    ws.cell(3, 14).value = "Проверка знаний Крановщики"
    ws.cell(3, 16).value = "Проверка знаний Люлька"
    ws.cell(3, 18).value = "Проверка знаний Высота"
    ws.cell(3, 20).value = "Противопожарный инструктаж"
    ws.cell(3, 22).value = "Инструктаж по ОТ"
    ws.cell(5, 2).value = "Иванов И.И."
    ws.cell(5, 3).value = "Машинист тепловоза"
    for col in [7, 9, 11, 13, 15, 17, 19, 21, 23]:
        ws.cell(5, col).value = datetime(2026, 6, 1)

    result = KnowledgeParser().parse(wb)

    assert result.rows_found == 9
    assert {row["check_type"] for row in result.rows} == {
        "Фильтр",
        "Проверка знаний по ОТ",
        "Проверка знаний по ЭБ",
        "Проверка знаний Стропальщики",
        "Проверка знаний Крановщики",
        "Проверка знаний Люлька",
        "Проверка знаний Высота",
        "Противопожарный инструктаж",
        "Инструктаж по ОТ",
    }


def test_screen_parser_handles_tab_number_column_shift() -> None:
    wb = Workbook()
    ws = wb.active
    ws.cell(3, 2).value = "ФИО"
    ws.cell(3, 3).value = "Профессия (должность)"
    ws.cell(3, 4).value = "Табельный номер"
    ws.cell(3, 5).value = "Мед. комиссия"
    ws.cell(3, 7).value = "Фильтр"
    ws.cell(3, 9).value = "Проверка знаний по ОТ"
    ws.cell(4, 5).value = "Дата предыдущей"
    ws.cell(4, 6).value = "Дата следующей"
    ws.cell(4, 7).value = "Дата предыдущего"
    ws.cell(4, 8).value = "Дата следующего"
    ws.cell(4, 9).value = "Дата предыдущей"
    ws.cell(4, 10).value = "Дата следующей"
    ws.cell(5, 2).value = "Иванов И.И."
    ws.cell(5, 3).value = "Машинист тепловоза"
    ws.cell(5, 4).value = 1001449
    ws.cell(5, 5).value = datetime(2025, 11, 7)
    ws.cell(5, 6).value = datetime(2026, 11, 7)
    ws.cell(5, 7).value = datetime(2025, 9, 9)
    ws.cell(5, 8).value = datetime(2028, 9, 9)
    ws.cell(5, 9).value = datetime(2025, 11, 19)
    ws.cell(5, 10).value = datetime(2026, 11, 17)

    medical = MedicalParser().parse(wb)
    knowledge = KnowledgeParser().parse(wb)

    assert medical.rows_found == 1
    assert medical.rows[0]["previous_date"].isoformat() == "2025-11-07"
    assert medical.rows[0]["next_date"].isoformat() == "2026-11-07"
    assert medical.rows[0]["tab_number"] == "1449"
    assert knowledge.rows_found == 2
    assert [(row["check_type"], row["previous_date"].isoformat(), row["next_date"].isoformat()) for row in knowledge.rows] == [
        ("Фильтр", "2025-09-09", "2028-09-09"),
        ("Проверка знаний по ОТ", "2025-11-19", "2026-11-17"),
    ]
