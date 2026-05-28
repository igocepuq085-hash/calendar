from datetime import date, datetime

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker

from app.database import Base
from app.importer import confirm_import, create_uploaded_file, detect_and_parse
from app.models import Employee, EmployeeStatus, KnowledgeCheck, MedicalCheck, ShiftType, WorkShift
from app.services.kip import default_shift_times


@pytest.fixture()
def db() -> Session:
    engine = create_engine("sqlite:///:memory:", connect_args={"check_same_thread": False})
    Base.metadata.create_all(engine)
    SessionLocal = sessionmaker(bind=engine, expire_on_commit=False)
    session = SessionLocal()
    try:
        yield session
    finally:
        session.close()


def _screen_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.cell(3, 2).value = "ФИО"
    ws.cell(3, 3).value = "Профессия (должность)"
    ws.cell(3, 4).value = "Табельный номер"
    ws.cell(3, 5).value = "Мед. комиссия"
    ws.cell(3, 7).value = "Проверка знаний по ОТ"
    ws.cell(4, 5).value = "Дата предыдущей"
    ws.cell(4, 6).value = "Дата следующей"
    ws.cell(4, 7).value = "Дата предыдущей"
    ws.cell(4, 8).value = "Дата следующей"
    return wb


def _save_upload(db: Session, tmp_path, wb: Workbook):
    path = tmp_path / "screen.xlsx"
    wb.save(path)
    result = detect_and_parse(path)
    return create_uploaded_file(db, original_filename="screen.xlsx", stored_path=str(path), result=result)


def test_screen_import_removes_medical_date_absent_from_new_file(db: Session, tmp_path) -> None:
    worker = Employee(full_name="Батятин В.И.", tab_number=None, department="СЛХ", status=EmployeeStatus.active)
    stale = Employee(full_name="Романенко О.А.", tab_number=None, department="битая строка", status=EmployeeStatus.active)
    archived = Employee(full_name="Архивный И.И.", tab_number="777", department="СЛХ", status=EmployeeStatus.inactive)
    db.add_all([worker, stale, archived])
    db.flush()
    db.add(MedicalCheck(employee_id=worker.id, previous_date=date(2024, 12, 27), next_date=date(2025, 12, 27)))
    db.add(KnowledgeCheck(employee_id=worker.id, check_type="Проверка знаний Высота", previous_date=date(2023, 9, 6), next_date=date(2024, 9, 4)))
    db.flush()

    wb = _screen_workbook()
    ws = wb.active
    ws.cell(5, 2).value = "Батятин В.И."
    ws.cell(5, 3).value = "Слесарь по ОРЛ на ПТО"
    ws.cell(5, 4).value = 1000868
    ws.cell(5, 7).value = datetime(2026, 2, 4)
    ws.cell(5, 8).value = datetime(2027, 2, 2)
    ws.cell(6, 2).value = "Козлов А.А."
    ws.cell(6, 3).value = "Машинист тепловоза"
    ws.cell(6, 4).value = 1001508
    ws.cell(6, 5).value = datetime(2026, 2, 24)
    ws.cell(6, 6).value = datetime(2027, 2, 24)
    ws.cell(7, 2).value = "Архивный И.И."
    ws.cell(7, 3).value = "Машинист тепловоза"
    ws.cell(7, 4).value = 1000777
    ws.cell(7, 7).value = datetime(2026, 2, 4)
    ws.cell(7, 8).value = datetime(2027, 2, 2)

    upload = _save_upload(db, tmp_path, wb)
    confirm_import(db, upload)

    batyatin = db.scalar(select(Employee).where(Employee.full_name == "Батятин В.И."))
    assert batyatin is not None
    assert batyatin.tab_number == "868"
    assert db.scalars(select(MedicalCheck).where(MedicalCheck.employee_id == batyatin.id)).all() == []
    assert [(row.check_type, row.next_date.isoformat()) for row in db.scalars(select(KnowledgeCheck).where(KnowledgeCheck.employee_id == batyatin.id))] == [
        ("Проверка знаний по ОТ", "2027-02-02")
    ]
    assert db.get(Employee, stale.id).status == EmployeeStatus.inactive
    assert db.get(Employee, archived.id).status == EmployeeStatus.inactive


def test_screen_import_merges_name_duplicate_by_tab_number(db: Session, tmp_path) -> None:
    name_row = Employee(full_name="Стругов А.Н.", tab_number=None, department="СЛХ", status=EmployeeStatus.active)
    tab_row = Employee(full_name="Стругов А.В.", tab_number="1650", department="СЛХ", status=EmployeeStatus.active)
    db.add_all([name_row, tab_row])
    db.flush()
    start, end = default_shift_times(date(2026, 5, 10), ShiftType.day)
    db.add(
        WorkShift(
            employee_id=tab_row.id,
            shift_date=date(2026, 5, 10),
            shift_type=ShiftType.day,
            start_datetime=start,
            end_datetime=end,
            raw_value="11.5",
        )
    )
    db.flush()

    wb = _screen_workbook()
    ws = wb.active
    ws.cell(5, 2).value = "Стругов А.Н."
    ws.cell(5, 3).value = "Машинист тепловоза"
    ws.cell(5, 4).value = 1001650
    ws.cell(5, 5).value = datetime(2025, 11, 18)
    ws.cell(5, 6).value = datetime(2026, 11, 18)
    ws.cell(5, 7).value = datetime(2025, 12, 17)
    ws.cell(5, 8).value = datetime(2026, 12, 15)

    upload = _save_upload(db, tmp_path, wb)
    confirm_import(db, upload)

    strugov_rows = db.scalars(select(Employee).where(Employee.full_name.contains("Стругов"))).all()
    assert [(row.full_name, row.tab_number) for row in strugov_rows] == [("Стругов А.Н.", "1650")]
    strugov = strugov_rows[0]
    assert db.scalar(select(WorkShift).where(WorkShift.employee_id == strugov.id)) is not None
    assert db.scalar(select(MedicalCheck).where(MedicalCheck.employee_id == strugov.id)).next_date.isoformat() == "2026-11-18"
